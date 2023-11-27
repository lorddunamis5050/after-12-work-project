import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from openpyxl.styles import PatternFill

def perform_regular_pick_analysis(df , book):
    # Define your desired time range for REGULAR PICK
    START_TIME_REGULAR = pd.to_datetime('12:00 AM', format='%I:%M %p')
    END_TIME_REGULAR = pd.to_datetime('6:30 AM', format='%I:%M %p')


    df['DateTime'] = pd.to_datetime(df['DateTime'])

    # Initialize DataFrames to store regular pick data
    regular_pick_per_user = pd.DataFrame(columns=['UserID', 'RegularPickQuantity'])

    # Function to modify the 'Action' column based on 'BinLabel' for REGULAR PICK
    def modify_action_regular(row):
        action = row['Action']
        bin_label = row['BinLabel']
        packslip = row['Packslip']

        if action == 'PICKLINE':
            if bin_label.startswith(('1H', '1G', '2E', '2H', '3F', '3H', '3R', '2R', '1Y', '1C', '1D', '2D', '3D')) and not packslip.startswith('TR'):
                return 'REGULAR PICK'

        return action

    # Apply the function to the DataFrame for REGULAR PICK
    df['Action'] = df.apply(modify_action_regular, axis=1)

    # Filter rows based on the specified time range for REGULAR PICK
    filtered_df_regular = df[(df['DateTime'] >= START_TIME_REGULAR) & (df['DateTime'] <= END_TIME_REGULAR)]


        # Group by "UserID" and calculate total Units picked
    regular_pick_per_user = filtered_df_regular[filtered_df_regular['Action'] == 'REGULAR PICK'].groupby('UserID').agg(
        RegularPickQuantity=('Quantity', 'sum')
    ).reset_index()

    def calculate_regular_picking_time(group):
        group = group.sort_values('DateTime')
        
        # Calculate the duration in minutes to the next row
        # No need to shift the 'DateTime' column before calculating the difference
        group['Duration'] = group['DateTime'].diff().dt.total_seconds().div(60).abs()

        # Mark the rows that are not part of continuous putwall picking
        # Also consider if the duration to the next action is more than 10 minutes as a gap
        group['IsGap'] = (~group['Action'].eq('REGULAR PICK') |
                        group['Action'].shift().ne('REGULAR PICK') |
                        (group['Duration'] > 10))

        # Cumulatively sum the gap marks to create unique session IDs
        group['SessionId'] = group['IsGap'].cumsum()

        # Filter out non-putwall picking rows and rows that start a new session
        regular_sessions = group[group['Action'].eq('REGULAR PICK') & ~group['IsGap']]

        # Calculate start and end time for each session
        session_times = regular_sessions.groupby('SessionId').agg({'DateTime': ['min', 'max']}) 

        # Calculate duration for each session
        session_durations = (session_times['DateTime']['max']  - session_times['DateTime']['min']).dt.total_seconds().div(60) 

        return session_durations.sum()  # Return the total time across all sessions
    
        # Filter for regular pick actions only
    regular_picking_actions = filtered_df_regular[filtered_df_regular['Action'] == 'REGULAR PICK']

        # Calculate total putwall regular time for each user considering gaps
    total_regular_picking_time = regular_picking_actions.groupby('UserID').apply(calculate_regular_picking_time).reset_index(name='Time')

        # Merge this time with the regular_picking_per_user DataFrame
    regular_pick_per_user = regular_pick_per_user.merge(total_regular_picking_time, on='UserID', how='left')

    # Calculate UPH (Units Per Hour) for each user using the total regular picking time
    regular_pick_per_user['UPH'] = regular_pick_per_user.apply(
    lambda row: (row['RegularPickQuantity'] * 60) / row['Time'] if row['Time'] >= 30 else 0, axis=1
    )



        # Convert both "RegularPickingQuantity" and "UPH" values to their absolute values
    regular_pick_per_user['RegularPickQuantity'] = abs(regular_pick_per_user['RegularPickQuantity'])
    regular_pick_per_user['UPH'] = abs(regular_pick_per_user['UPH']).round(2)

        # Calculate the average UPH, excluding zeros
    average_uph = regular_pick_per_user.loc[regular_pick_per_user['UPH'] > 0, 'UPH'].mean().round(2)

            # Replace NaN or infinite values with zero
    regular_pick_per_user.replace([np.inf, -np.inf, np.nan], 0, inplace=True)

    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")



        # Create the "REGULAR PICKING" sheet if it doesn't exist
    if 'REGULAR PICK' not in book.sheetnames:
        regular_pick_sheet = book.create_sheet('REGULAR PICK')
    else:
        regular_pick_sheet = book['REGULAR PICK']

            # Write the header row
    header_row = ['UserID', 'RegularPickQuantity','Time', 'UPH']
    regular_pick_sheet.append(header_row)

            # Format header row with light blue background
    for cell in regular_pick_sheet[1]:
        cell.fill = light_blue_fill

        # Sort by UPH in descending order
    regular_pick_per_user.sort_values(by='UPH', ascending=False, inplace=True)

            # Calculate the average UPH, excluding zeros, and round to 2 decimal places
    average_uph = regular_pick_per_user.loc[regular_pick_per_user['UPH'] > 0, 'UPH'].mean().round(2)


        # Convert the DataFrame to a list of lists for writing to Excel
    regular_picking_data = regular_pick_per_user[['UserID', 'RegularPickQuantity','Time' ,'UPH']].values.tolist()

    # Write the data to the Excel sheet
    for row_data in regular_picking_data:
        regular_pick_sheet.append(row_data)

        # Write the average UPH row to the Excel sheet
    average_uph_row = ["Average UPH", "", "", average_uph]
    regular_pick_sheet.append(average_uph_row)

    for cell in regular_pick_sheet[regular_pick_sheet.max_row]:
        cell.fill = light_blue_fill    


    print("REGULAR PICKING analysis completed.")
