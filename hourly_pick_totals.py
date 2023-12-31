import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Define the time range
START_TIME = pd.to_datetime('12:00 AM', format='%I:%M %p')
END_TIME = pd.to_datetime('6:30 AM', format='%I:%M %p')

def calculate_pick_totals(df, action_filter):
    # Filter the DataFrame by action_filter and time range
    filtered_df = df[(df['Action'].isin(action_filter)) & (df['DateTime'].dt.hour >= 0) & (df['DateTime'].dt.hour < 7)]
    
    # Group by hour and sum quantities
    pick_totals = filtered_df.groupby(filtered_df['DateTime'].dt.hour)['Quantity'].apply(lambda x: abs(x).sum()).to_dict()
    
    return pick_totals

def perform_hourly_pick_totals_analysis(df, book):
    # Define the pick types and corresponding action filters
    pick_types = {
        'Regular Pick': ['REGULAR PICK'],
        'Single Pick': ['SINGLE PICK'],
        'Replenishment Pick': ['REPLENISHMENT PICK'],
        'Putwall Pick': ['PUTWALL PICKING']
    }

    # Create the "Total Units picked by hour" sheet if it doesn't exist
    if 'Total Units picked by hour' not in book.sheetnames:
        hourly_pick_totals_sheet = book.create_sheet('Total Units picked by hour')
    else:
        hourly_pick_totals_sheet = book['Total Units picked by hour']

    # Write the header row
    header_row = ['Hour'] + list(pick_types.keys())
    hourly_pick_totals_sheet.append(header_row)

    # Calculate and write the total quantity for each hour and pick type
    for hour in range(0, 7):
        hour_data = [hour]

        for pick_type, action_filter in pick_types.items():
            pick_totals = calculate_pick_totals(df, action_filter)
            quantity = pick_totals.get(hour, 0)
            hour_data.append(quantity)

        hourly_pick_totals_sheet.append(hour_data)

    # Calculate and write the total quantities for each pick type
    total_row = ['Total']
    for pick_type, action_filter in pick_types.items():
        pick_totals = calculate_pick_totals(df, action_filter)
        total_quantity = sum(pick_totals.values())
        total_row.append(total_quantity)

    hourly_pick_totals_sheet.append(total_row)

    # Save the Excel file
    output_excel_file = 'pick_counts.xlsx'
    book.save(output_excel_file)

    print("Hourly pick totals analysis completed and saved.")

# Usage:
# Call perform_hourly_pick_totals_analysis with your DataFrame and Excel book as arguments
# Example: perform_hourly_pick_totals_analysis(df, openpyxl.Workbook())
